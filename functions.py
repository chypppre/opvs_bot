import datetime
from openpyxl import load_workbook, Workbook
import sqlite3

from telebot import types

from misc import DB_DIR, REPORTS_DIR, CHATS_ID, OO_DIR

now = datetime.datetime.now()
current_day = now.day
current_hour = now.hour

# -----------------------ФУНКЦИИ ДЛЯ ДЕЛИШЕК----------------------------------


def print_in_console(text, date, fullname=""):
    """Функция для печати инф-ы с датой/временем в консоль"""
    print(str(date))
    if fullname != "":
        print("{} {}\n".format(text, fullname))
    else:
        print("{}\n".format(text))


def auth(bot, uid):
    """Проверить живет ли конс в БД"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM staff WHERE uid={}".format(uid)
    cursor.execute(sql)
    result = cursor.fetchall()
    if len(result) == 0:  # Если ответ на запрос пуст, то юзера нет в БД.
        bot.send_message(
                chat_id=uid,
                text="Похоже, что бот тебя не узнал. Напиши команду /addmeplease")
        return False
    else:
        return True


def is_intern(uid):
    """Это ли помогатор?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM staff WHERE department=3"
    cursor.execute(sql)
    interns_uid = cursor.fetchall()
    for intern in interns_uid:
        if uid == intern[0]:
            return True
    return False


def is_mentor(uid):
    """Это ли помогатор?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM mentors"
    cursor.execute(sql)
    mentors_uid = cursor.fetchall()
    for mentor in mentors_uid:
        if uid == mentor[0]:
            return True
    return False


def is_helper(uid):
    """Это ли помогатор?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM helpers"
    cursor.execute(sql)
    helpers_uid = cursor.fetchall()
    for helper in helpers_uid:
        if uid == helper[0]:
            return True
    return False


def is_super(uid):
    """Это ли супервизор?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM supers"
    cursor.execute(sql)
    supers_uid = cursor.fetchall()
    for _super in supers_uid:
        if uid == _super[0]:
            return True
    return False


def is_pip(uid):
    """Это ли ПиП?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM pips"
    cursor.execute(sql)
    pips_uid = cursor.fetchall()
    for pip in pips_uid:
        if uid == pip[0]:
            return True
    return False


def is_oo(uid):
    """Это ли из ОО?"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT uid FROM staff WHERE department=4"
    cursor.execute(sql)
    oos_uid = cursor.fetchall()
    for oo in oos_uid:
        if uid == oo[0]:
            return True
    return False


def create_keyboard(department):
    kbrd = types.ReplyKeyboardMarkup(resize_keyboard=True)
    ping_btn = types.KeyboardButton(text="Пинг")
    dezh_btn = types.KeyboardButton(text="Дежурный")
    #change_btn = types.KeyboardButton(text="Обратная связь")
    etc_btn = types.KeyboardButton(text="Прочие")
    oo_btn = types.KeyboardButton(text="ОО")
    if department == 1:
        kbrd.row(ping_btn, dezh_btn)
        kbrd.row(etc_btn)#, change_btn)
        return kbrd
    elif department == 3:
        kbrd.row(ping_btn, dezh_btn)
        kbrd.row(oo_btn)
        return kbrd
    elif department == "cancel":
        kbrd = types.InlineKeyboardMarkup()
        cancel_btn = types.InlineKeyboardButton(
            text="Отменить пинг.", callback_data="cancel")
        kbrd.row(cancel_btn)
    elif department == "dezh_cancel":
        kbrd = types.InlineKeyboardMarkup()
        cancel_btn = types.InlineKeyboardButton(
            text="Отменить дежурного.", callback_data="dezh_cancel")
        kbrd.row(cancel_btn)
    return kbrd


def get_department(where, uid):
    """Найти условный отдел консультанта"""
    db = sqlite3.connect(DB_DIR)
    cursor = db.cursor()
    sql = "SELECT department FROM {} WHERE uid={}".format(where, uid)
    data = cursor.execute(sql)
    return data.fetchall()[0][0]


def get_admin_chat_id(where, uid):
    """Найти id чата , в котором нужно редактировать сообщение"""
    department = get_department(where, uid)
    if department == 1:
        return CHATS_ID[0]
    elif department == 3:
        return CHATS_ID[4]


def send_members_list(bot, user_id, message, sql, m_text):
    """Отправить список пользователей согласно запросу"""
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    try:
        mycursor.execute(sql)
    except:
        text4 = "Список пуст!"
    member_list = mycursor.fetchall()
    e_text = []
    for row in member_list:
        if len(row) < 4:
            uid, last_name, first_name, *etc = row
            try:
                r_text = "{} {} {}".format(uid, last_name, first_name)
                e_text.append(r_text)
            except:
                pass
        else:
            uid, department, last_name, first_name, *etc = row
            try:
                r_text = "{} {} {} {}".format(uid, department, last_name, first_name)
                e_text.append(r_text)
            except:
                pass
    text4 = "\n".join(e_text)
    bot.edit_message_text(
            chat_id=user_id,
            message_id=message,
            text='{} {}'.format(m_text, text4))


def update_calls(what, department):
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    today = datetime.date.today()
    sql1 = "UPDATE stats SET date='{}-{:0>2}-{:0>2}' WHERE department={}".format(
            today.year, today.month, today.day, department)
    mycursor.execute(sql1)
    sql2 = "SELECT {} FROM stats WHERE department={}".format(what, department)
    mycursor.execute(sql2)
    result = mycursor.fetchall()
    value = result[0][0] + 1
    sql4 = "UPDATE stats SET {}={} WHERE department={}".format(what, value, department)
    mycursor.execute(sql4)
    mydb.commit()


def update_answers(uid, what, chat):
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    if is_mentor(uid) and chat == CHATS_ID[4]:
        sql = "SELECT {} FROM mentors WHERE uid={}".format(what, uid)
        mycursor.execute(sql)
        result = mycursor.fetchall()
        value = result[0][0] + 1
        sql4 = "UPDATE mentors SET {}={} WHERE uid={}".format(what, value, uid)
        mycursor.execute(sql4)
    elif is_helper(uid):
        sql = "SELECT {} FROM helpers WHERE uid={}".format(what, uid)
        mycursor.execute(sql)
        result = mycursor.fetchall()
        value = result[0][0] + 1
        sql4 = "UPDATE helpers SET {}={} WHERE uid={}".format(what, value, uid)
        mycursor.execute(sql4)
    elif is_super(uid):
        sql = "SELECT {} FROM supers WHERE uid={}".format(what, uid)
        mycursor.execute(sql)
        result = mycursor.fetchall()
        value = result[0][0] + 1
        sql2 = "UPDATE supers SET {}={} WHERE uid={}".format(what, value, uid)
        mycursor.execute(sql2)
    mydb.commit()
    mycursor.close()
    mydb.close()


def update_ping_status(uid, m_id, ping_type, ping_status, is_reply='False'):
    """Обновить инфу о состоянии пинга"""
    now = datetime.datetime.now()
    hours = now.hour
    minutes = now.minute
    seconds = now.second
    total_time = hours*60*60 + minutes*60 + seconds

    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()

    if not is_reply:
        sql = """UPDATE staff
                SET last_ping_mid={}, ping_type='{}', ping_status={}, time_to_answer={}
                WHERE uid={}""".format(m_id, ping_type, ping_status, total_time, uid)
    else:
        sql1 = """SELECT time_to_answer, total_time, department
                FROM staff
                WHERE uid={}""".format(uid)
        first_time, staff_t_time, department = mycursor.execute(sql1).fetchall()[0]
        sql = """SELECT total_time
                FROM stats
                WHERE department={}""".format(department)
        stats_t_time = mycursor.execute(sql).fetchall()[0][0]
        sql = """UPDATE staff
                SET last_ping_mid={}, ping_type='{}', ping_status={}, time_to_answer=0, total_time={}
                WHERE uid={}""".format(m_id, ping_type, ping_status, staff_t_time+(total_time-first_time), uid)
        mycursor.execute(sql)
        sql = """UPDATE stats
                SET total_time={}
                WHERE department={}""".format(stats_t_time+(total_time-first_time), department)
    mycursor.execute(sql)

    if is_reply == False:
        value = 1
    else:
        value = 0

    if ping_type != 'empty':
        sql = "SELECT {} FROM staff WHERE uid={}".format(ping_type, uid)
        res = mycursor.execute(sql)
        data = res.fetchall()[0][0] + value
        sql = "UPDATE staff SET {}={} WHERE uid={}".format(ping_type, data, uid)
    else:
        sql = "UPDATE staff SET ping_type='empty' WHERE uid={}".format(uid)
    mycursor.execute(sql)
    mydb.commit()


def update_dezh_status(uid, last_dezh_mid, dezh_type, dezh_status, is_reply='False'):
    """Обновить инфу о состоянии пинга"""
    now = datetime.datetime.now()
    hours = now.hour
    minutes = now.minute
    seconds = now.second
    total_time = hours*60*60 + minutes*60 + seconds

    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()

    if not is_reply:
        sql = "UPDATE staff SET dezh_status={}, last_dezh_mid={}, time_to_answer_dezh={} WHERE uid={}".format(
                dezh_status, last_dezh_mid, total_time, uid)
    else:
        sql1 = """SELECT time_to_answer_dezh, total_dezh_time, department
                FROM staff
                WHERE uid={}""".format(uid)
        first_time, staff_t_time, department = mycursor.execute(sql1).fetchall()[0]
        sql = """SELECT total_time
                FROM stats
                WHERE department={}""".format(department)
        stats_t_time = mycursor.execute(sql).fetchall()[0][0]
        sql = """UPDATE staff
                SET dezh_status={}, last_dezh_mid={}, time_to_answer_dezh=0, total_dezh_time={}
                WHERE uid={}""".format(dezh_status, last_dezh_mid, staff_t_time+(total_time-first_time), uid)
        mycursor.execute(sql)
        sql = """UPDATE stats
                SET total_dezh_time={}
                WHERE department={}""".format(stats_t_time+(total_time-first_time), department)
    mycursor.execute(sql)

    if is_reply == False:
        value = 1
    else:
        value = 0

    if dezh_type != 'empty':
        sql = "SELECT {} FROM staff WHERE uid={}".format(dezh_type, uid)
        mycursor.execute(sql)
        data = mycursor.fetchall()[0][0]
        amt = data + value
        sql2 = "UPDATE staff SET {}={}, dezh_status={}, last_dezh_mid={} WHERE uid={}".format(
            dezh_type, amt, dezh_status, last_dezh_mid, uid)
    else:
        sql2 = "UPDATE staff SET dezh_status={}, last_dezh_mid={} WHERE uid={}".format(
            dezh_status, last_dezh_mid, uid)
    mycursor.execute(sql2)
    mydb.commit()


def mass_message(bot, text, uid):
    """Отправить сообщение всем пользователям бота"""
    text = text[text.find("\"")+1 : text.rfind("\"")]
    users = set()
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    sql = """SELECT uid
            FROM staff"""
    mycursor.execute(sql)
    result = mycursor.fetchall()
    for user in result:
        users.add(user[0])
    for uid in users:
        try:
            bot.send_message(
                    chat_id=uid,
                    text=text)
        except:
            bot.send_message(
                    chat_id=uid,
                    text="для {} не отправилось".format(uid))
    bot.send_message(
            chat_id=uid,
            text="Все получили сообщение.")


def group_message(bot, text, uid):
    """Отправить сообщение группе пользователей"""
    group = text.split(" ")[1]
    message = text[text.find("\"")+1 : text.rfind("\"")]
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    users = set()

    if group == "экстерн":
        sql = "SELECT uid FROM staff WHERE department=1"
        for user in mycursor.execute(sql).fetchall():
            users.add(user[0])
        try:
            for user in users:
                bot.send_message(
                    chat_id=user,
                    text=message)
        except:
            bot.send_message(
                chat_id=uid,
                text="Не доставлено {}".format(user))

    elif group == "стажеры":
        sql = "SELECT uid FROM staff WHERE department=3"
        for user in mycursor.execute(sql).fetchall():
            users.add(user[0])
        try:
            for user in users:
                bot.send_message(
                    chat_id=user,
                    text=message)
        except:
            bot.send_message(
                chat_id=uid,
                text="Не доставлено {}".format(user))

    elif group == "помогаторы":
        sql = "SELECT uid FROM helpers"
        for user in mycursor.execute(sql).fetchall():
            users.add(user[0])
        try:
            for user in users:
                bot.send_message(
                    chat_id=user,
                    text=message)
        except:
            bot.send_message(
                chat_id=uid,
                text="Не доставлено {}".format(user))

    elif group == "пипы":
        sql = "SELECT uid FROM pips"
        for user in mycursor.execute(sql).fetchall():
            users.add(user[0])
        try:
            for user in users:
                bot.send_message(
                    chat_id=user,
                    text=message)
        except:
            bot.send_message(
                chat_id=uid,
                text="Не доставлено {}".format(user))

    elif group == "суперыкэ":
        sql = "SELECT uid FROM supers WHERE department=1"
        for user in mycursor.execute(sql).fetchall():
            users.add(user[0])
        try:
            for user in users:
                bot.send_message(
                    chat_id=user,
                    text=message)
        except:
            bot.send_message(
                chat_id=uid,
                text="Не доставлено {}".format(user))

    bot.send_message(
            chat_id=uid,
            text="Все получили сообщение.")


def private_message(bot, text, uid):
    """Отправить сообщение пользователю бота"""
    try:
        uid = text.split(" ")[1]
        msg = text[text.find("\"")+1:-1]
        bot.send_message(
                chat_id=uid,
                text=msg)
    except:
        print("OOOOOOOOOOOOOOOOOOOOOPS!!!")


def status(bot, adress, uid):
    """Поменять адрес пребывания"""
    mydb = sqlite3.connect(DB_DIR)
    cursor = mydb.cursor()
    # Супервизорский
    if adress.startswith("Установить супервизорский адрес"):
        adress = adress[len('Установить супервизорский адрес '):]
        sql = """UPDATE supers
                SET adress='{}'
                WHERE uid={}""".format(adress, uid)
    # Помогаторский
    elif adress.startswith("Установить помогаторский адрес"):
        adress = adress[len('Установить помогаторский адрес '):]
        sql = """UPDATE helpers
                SET adress='{}'
                WHERE uid={}""".format(adress, uid)
    # Консультантский
    elif adress.startswith("Установить адрес"):
        adress = adress[len('Установить адрес '):]
        sql = """UPDATE staff
                SET adress='{}'
                WHERE uid={}""".format(adress, uid)
    # Установить имя
    elif adress.startswith("Установить имя"):
        adress = adress[len('Установить имя '):]
        sql = """UPDATE staff
                SET first_name='{}'
                WHERE uid={}""".format(adress, uid)
    # Установить фамилию
    elif adress.startswith("Установить фамилия"):
        adress = adress[len('Установить фамилия '):]
        sql = """UPDATE staff
                SET last_name='{}'
                WHERE uid={}""".format(adress, uid)
    cursor.execute(sql)
    mydb.commit()
    bot.send_message(
            chat_id=uid,
            text="Адрес изменен на {}. Теперь напиши боту /start.".format(adress))
    cursor.close()


def oo_info(bot, message):
    department = get_department('staff', message.from_user.id)

    if message.text in ["ОО", "Меню ОО"]:
        oo_kbrd = types.ReplyKeyboardMarkup(resize_keyboard=True)
        reg_btn = types.KeyboardButton(text="Регламенты")
        staff_btn = types.KeyboardButton(text="Сотрудники ОО")
        wtf_btn = types.KeyboardButton(text="Что делать, если...")
        cont_btn = types.KeyboardButton(text="Контакты")
        back_btn = types.KeyboardButton(text="Главное меню")
        oo_kbrd.row(reg_btn, staff_btn)
        oo_kbrd.row(wtf_btn, cont_btn)
        oo_kbrd.row(back_btn)
        bot.send_message(
                chat_id=message.from_user.id,
                text="Отдел Обучения",
                reply_markup=oo_kbrd)

    elif message.text == "Главное меню":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Главное меню.",
                reply_markup=create_keyboard(department))

    elif message.text == "Регламенты":
        reg_kbrd = types.ReplyKeyboardMarkup(resize_keyboard=True)
        codes_btn = types.KeyboardButton(text="Коды перерывов для Yealink")
        vw_btn = types.KeyboardButton(text="Very Want и Want")
        codes2_btn = types.KeyboardButton(text="Для чего какой код")
        money_btn = types.KeyboardButton(text="Показатели премии")
        money2_btn = types.KeyboardButton(text="Дни зарплаты")
        back_btn = types.KeyboardButton(text="Главное меню")
        oo_menu_btn = types.KeyboardButton(text="Меню ОО")
        reg_kbrd.row(codes_btn, vw_btn)
        reg_kbrd.row(codes2_btn, money_btn)
        reg_kbrd.row(money2_btn)
        reg_kbrd.row(back_btn, oo_menu_btn)
        bot.send_message(
                chat_id=message.from_user.id,
                text="Регламенты.",
                reply_markup=reg_kbrd)

    elif message.text == "Коды перерывов для Yealink":
        try:
            with open(OO_DIR+"yealink.jpg", 'rb') as file:
                bot.send_document(message.from_user.id, file)
        except:
            pass
        try:
            with open(OO_DIR+"yealink.pdf", 'rb') as file:
                bot.send_document(message.from_user.id, file)
        except:
            pass

    elif message.text == "Very Want и Want":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Кнопка «WANT» для перерыва до 60 мин."\
                    "\nНе стоит тратить его за раз или задерживаться — твои коллеги с тобой в общей очереди."\
                    "\n\nКнопка «VERY WANT» для перерыва на 10 мин."\
                    "\n\nПосле того, как ты нажал «GO» на callider, поменяй тип перерыва на телефоне на «Личный».")

    elif message.text == "Для чего какой код":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Другое поручение — поручения наставников и супервизоров."\
                    "\n\nПроверка знаний — повышение квалификации."\
                    "\n\nБеседа с супервизором —  ОС по качеству твоей работы."\
                    "\n\nПочта — только для писем пользователей."\
                    "\n\nЧаты — только для чатов."\
                    "\n\nАльтработа — чтобы анализировать обращения от пользователей, собирать статистику."\
                    "\n\nЛичный — для перерывов на обед или отдыха."\
                    "\n\nУдаленный доступ — для подключения к ПК пользователя."\
                    "\n\nОбучение — лекции, тренинги и самопрослушивание."\
                    "\n\nИсходящий вызов — чтобы позвонить пользователю."\
                    "\n\nСложный инцидент — чтобы заполнить инцидент для экспертов. "\
                    "В этом коде нужно указывать номер инцидента."\
                    "\n\nБлиц-опрос — тесты на знание материала и регламентов."\
                    "\n\nТайм-аут с клиентом — чтобы подготовить ответ для клиента, который не хочет ожидать на линии."\
                    "В этом коде нужно указывать номер инцидента.")

    elif message.text == "Показатели премии":
        bot.send_message(
                chat_id=message.from_user.id,
                text="После перехода на трудовой договор ты можешь получать премию в 20% каждый месяц."\
                    "\n\nПремия зависит от:"\
                    "\n1. Качества консультаций."\
                    "\n2. Процента рассказанных уведомлений и зарегистрированных обращений."\
                    "\n3. Целевая занятость."\
                    "\n\nПодробнее о расчёте премии читай на "\
                    "wiki https://wiki.skbkontur.ru/pages/viewpage.action?pageId=307512810")

    elif message.text == "Дни зарплаты":
        bot.send_message(
                chat_id=message.from_user.id,
                text="В день заключения трудового договора тебе придет стипендия."\
                    "\n\nДень зарплаты — 10 число каждого месяца. Аванс— 25 число.")

    elif message.text == "Сотрудники ОО":
        staff_kbrd = types.ReplyKeyboardMarkup(resize_keyboard=True)
        head_btn = types.KeyboardButton(text="Руководители")
        #ekb_btn = types.KeyboardButton(text="Екатеринбург")
        #vlg_btn = types.KeyboardButton(text="Волгоград")
        #novosib_btn = types.KeyboardButton(text="Новосибирск")
        vor_btn = types.KeyboardButton(text="Воронеж")
        trainers_btn = types.KeyboardButton(text="Тренеры")
        back_btn = types.KeyboardButton(text="Главное меню")
        oo_menu_btn = types.KeyboardButton(text="Меню ОО")
        staff_kbrd.row(head_btn)
        staff_kbrd.row(ekb_btn, vlg_btn)
        staff_kbrd.row(novosib_btn, vor_btn)
        staff_kbrd.row(trainers_btn)
        staff_kbrd.row(back_btn, oo_menu_btn)
        bot.send_message(
                chat_id=message.from_user.id,
                text="Сотрудники ОО.",
                reply_markup=staff_kbrd)

    elif message.text == "Руководители":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Третьякова Екатерина — руководитель отдела обучения."\
                    "\n@trekaterina, +7 909 018-25-21.")

    #elif message.text == "Екатеринбург":
    #    bot.send_message(
    #            chat_id=message.from_user.id,
    #            text="❇️ Алексеев Алексей @lexxxekb"\
    #                "\nрассказывает как работать с телефоном, уведомлениями, знаёт всё про сервисы УТО и сертификаты."\
    #                "\n\n❇️ Баталина Ирина @irinabatalina"\
    #                "\nзнает всё про сертификаты, переводит лекции в онлайн."\
    #                "\n\n❇️ Брюханов Никита @nikitabrukhanov"\
    #                "\nчитает лекции по сервисам АУБ и стилю текстов."\
    #                "\n\n❇️ Вязовиков Константин @vyazovikov"\
    #                "\nсоставляет план работы ОО, читает лекции для ОПВС."\
    #                "\n\n❇️ Григорьев Иван @Pyatackovski"\
    #                "\nзнает всё про сертификаты, читает лекции по сервисам ЭП."\
    #                "\n\n❇️ Иванова Ольга @ivanovaom"\
    #                "\nзнает всё про Маркет, читает лекции по сервисам УТО. "\
    #                "Специалист по написанию статей, инструкций и просто красивых текстов!"\
    #                "\n\n❇️ Криницына Анна @anna_krinitsyna"\
    #                "\nзанимается планами на каждый день, знает всё про Экстерн."\
    #                "\n\n❇️ Кудряшов Михаил @Greenz42"\
    #                "\nчитает лекции для АРТ, рассказывает о кассах, сертификатах и настройке ПК."\
    #                "\n\n❇️ Неркарарян Армен @ArmenNerkararyan"\
    #                "\nчитает лекции по КЭ и настройке рабочего места, знает всё про внутренние ресурсы."\
    #                "\n\n❇️ Протасенко Кристина @protasenko_kristina"\
    #                "\nчитает лекции по сервисам ЭДО."\
    #                "\n\n❇️ Русинова Александра @allergique"\
    #                "\nчитает лекции по внутренним ресурсам, знает всё про НДС и требования."\
    #                "\n\n❇️ Рязанов Михаил @Mishuk"\
    #                "\nчитает лекции по сервисам АУБ."\
    #                "\n\n❇️ Савченко Алёна @savchenko_alena"\
    #                "\nв декрете.")

    #elif message.text == "Волгоград":
    #    bot.send_message(
    #            chat_id=message.from_user.id,
    #            text="❇️ Дьяконова Валерия @Lera_Dyakonova"\
    #                "\nчитает лекции по сервисам ЭДО и ОПВС, знает всё об информационной безопасности."\
    #                "\n\n❇️ Погодина Марина @m_bagira"\
    #                "\nчитает лекции по сервисам ОПВС.")

    #elif message.text == "Новосибирск":
    #    bot.send_message(
    #            chat_id=message.from_user.id,
    #            text="❇️ Хагай Ксения @Kusyuma1"\
    #                "\nобучает стажёров сервисам УТО, помогает вести тренинги и отвечает на вопросы пользователей."\
    #                "\n\n❇️ Кучер Артем @flartman"\
    #                "\nстажируется в отделе обучения, помогает пользователям Экстерна и Фокуса.")

    elif message.text == "Воронеж":
        bot.send_message(
                chat_id=message.from_user.id,
                text="❇️ Новикова Марина @m_malkova"\
                    "\nчитает лекции по сервисам ОПВС."\
                    "\n\n❇️ Зотьев Роман @MarcusAvrelius"\
                    "\nчитает лекции по сервисам ОПВС.")

    elif message.text == "Тренеры":
        bot.send_message(
                chat_id=message.from_user.id,
                text="❇️ Грефенштейн Петр @PeterGref"\
                    "\n\n❇️ Сергеева Дарья @Dasha_Sergeeva"
                    "\n\n❇️ Черных Людмила @lyukache")

    elif message.text == "Что делать, если...":
        wtf_kbrd = types.ReplyKeyboardMarkup(resize_keyboard=True)
        sick_btn = types.KeyboardButton(text="Я заболел")
        late_btn = types.KeyboardButton(text="Я опаздываю")
        otgul_btn = types.KeyboardButton(text="Нужен отгул")
    #    hungry_btn = types.KeyboardButton(text="Проголодался")
        blood_btn = types.KeyboardButton(text="Я сдаю кровь")
        army_btn = types.KeyboardButton(text="Пришла повестка")
        insurance_btn = types.KeyboardButton(text="Нужна медицинская страховка")
        out_btn = types.KeyboardButton(text="Надо отпроситься")
        back_btn = types.KeyboardButton(text="Главное меню")
        oo_menu_btn = types.KeyboardButton(text="Меню ОО")
        wtf_kbrd.row(sick_btn, late_btn)
        wtf_kbrd.row(otgul_btn, hungry_btn)
        wtf_kbrd.row(blood_btn, army_btn)
        wtf_kbrd.row(insurance_btn, out_btn)
        wtf_kbrd.row(back_btn, oo_menu_btn)
        bot.send_message(
                chat_id=message.from_user.id,
                text="Сотрудники ОО.",
                reply_markup=wtf_kbrd)

    elif message.text == "Я заболел":
        bot.send_message(
                chat_id=message.from_user.id,
                text="1. Напиши @m_malkova или наставнику, что болеешь. "\
                    "\n2. Посети врача и оформи больничный. "\
                    "Сообщи нам день начала больничного и день повторного приема: "\
                    "«Дали больничный 1.02, повторный приём 4.02. Иванов Иван»."\
                    "\n3. Сходи на повторный приём и напиши нам, когда больничный закроют. "\
                    "Если больничный продлили, напиши день повторного приёма.")

    elif message.text == "Я опаздываю":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Напиши @m_malkova или или наставнику, что опаздываешь. "\
                "В сообщении должны быть:"\
                "\n\n— ФИО"\
                "\n— дата набора"\
                "\n— причина опоздания"\
                "\n— время, когда появишься на работе.")

    elif message.text == "Нужен отгул":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Если с тобой ещё не заключили ТД, напиши @m_malkova."\
                    "\n\nЕсли с тобой заключили ТД, то:"\
                    "\n1. Договорись с наставником о днях, когда тебя не будет."\
                    "\n2. Отправь заявление на отпуск."\
                    "\n\nЧтобы написать заявление зайди на старый callider в раздел «График» "\
                    "→ «Заявления на отпуск». Выбери «Отпуск без сохранения заработной платы»."\
                    "\n\nВ поле для причины укажи «семейные обстоятельства» даже если реальная "\
                    "причина не такая."\
                    "\n\nЕсли отпуск нуже на 1 день, то в периоде укажи только этот день: "\
                    "если тебя не будет 8 июля, то период с 08.07.2019 по 08.07.2019."\
                    "\n\nЕсли заявление не получилось написать в день отпуска, то напиши его на "\
                    "следующий день в отделе персонала в 616.")

    #elif message.text == "Проголодался":
    #    bot.send_message(
    #            chat_id=message.from_user.id,
    #            text="Екатеринбург"\
    #                "\nПосле перехода на трудовой договор ты можешь заказывать обеды в офис на свой этаж. "\
    #                "Чтобы заказать обед, перейди на callider → «Еще» → «Обеды»."\
    #                "\n\nНа Радищева заказывать можно из Offline и Caterinburg."\
    #                "\n\nЗаказ на завтра надо сделать сегодня до 17:00. "\
    #                "Т. е., чтобы обед пришёл в понедельник, заказ надо сделать в воскресенье до 17:00."\
    #                "\n\nЕсли ты работаешь в Контур НТТ, то пока заказывать обеды нельзя, "\
    #                "но скоро такая возможность появится.")
    #    bot.send_message(
    #            chat_id=message.from_user.id,
    #            text="Новосибирск"\
    #                "\nПосле перехода на трудовой договор ты можешь заказывать обеды в офис."\
    #                "\n\nЧтобы заказать обед, перейди на obed.ru, авторизуйся. "\
    #                "Открой «Обеды» → «21 век». Заказ на завтра надо сделать в рабочий день до 14.00. "\
    #                "Т. е. чтобы обед пришел в понедельник, заказ надо сделать в пятницу до 14.00."\
    #                "\n\nРадом с офисом есть кафе «Оливка». Полный список кафе, подключенных к системе "\
    #                "Obed.ru можно посмотреть на сайте по кнопке «По карте Obed.ru» в правом верхнем углу.")

    elif message.text == "Я сдаю кровь":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Напиши @m_malkova, что собираешься сдать кровь. "\
                    "Если ты работаешь на ТД, то пиши наставнику."\
                    "\nЕсли ты выйдешь на работу в день, когда сдаёшь кровь, передай в отдел персонала "\
                    "фотографию или оригинал справки о сдаче крови в этот же день до 13:00. "\
                    "Работник отдела пресонала оформит приказ о допуске к работе в день сдачи крови и "\
                    "её компонентов."\
                    "\n\nЕсли в день сдачи крови ты не выйдешь на работу, напиши заявление об освобождении "\
                    "от работы в отделе персонала. Сделай этого заранее или утром в день, когда сдаёшь кровь."\
                    "\n\nЕсли ты сдаёшь кровь в выходной день или во время отпуска, "\
                    "то никакие документы оформлять не надо."\
                    "\n\nПосле ты можешь взять два дополнительных дня оплачиваемого отпуска. "\
                    "Для этого зайди на старый callider в раздел «График» → «Заявления на отпуск». "\
                    "Выбери «Дополнительный день отдыха за сдачу крови и ее компонентов».")

    elif message.text == "Пришла повестка":
        bot.send_message(
            chat_id=message.from_user.id,
            text="1. Напиши @m_malkova или наставнику, что тебе надо быть в военкомате. "\
                "\n2. Принеси в отдел персонала оригинал повестки. Копия не подойдёт."\
                "\n3. Там же напиши заявление об отсутствии на рабочем месте."\
                "\n\nВсё это надо сделать до дня визита в военкомат.")

    elif message.text == "Нужна медицинская страховка":
        bot.send_message(
                chat_id=message.from_user.id,
                text="К программе ДМС можно подключиться бесплатно, если твой стаж работы в компании 1 год. "\
                "В ином случае подключиться можно за свой счет.")

    elif message.text == "Надо отпроситься":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Напиши @m_malkova или наставнику, что тебе надо уйти раньше. "\
                "\n\nЕсли ты работаешь на ТД, напиши наставнику.")

    elif message.text == "Контакты":
        bot.send_message(
                chat_id=message.from_user.id,
                text="Пиши нам на OO_owner@skbkontur.ru"\
                "\n\nПиши коллегам в Telegram или СМС. Звони, если вопрос срочный. "\
                "Наши контакты по кнопке «Сотрудники ОО»."\
                "\n\nЕсли есть идеи для бота, пиши @fancyAndBeauty.")


def make_statistic(department):
    """Состаление отчета по пингам за сутки"""
    if department == 1:
        # путь до папки с отчетами
        PATH = REPORTS_DIR+"КЭ_{}_{}.xlsx".format(datetime.datetime.now().year, datetime.datetime.now().month)
        # перечисление нужных столбцов для отчета
        needed_losts = True
        needed_kcrs = True
        needed_exps = True
        needed_visions = True
        needed_pzs = True
        needed_miss = True
        needed_study = True
#    elif department == 2:
#        PATH = REPORTS_DIR+"ФМС_{}_{}.xlsx".format(datetime.datetime.now().year, datetime.datetime.now().month)
#        needed_losts = False
#        needed_kcrs = False
#        needed_exps = False
#        needed_visions = True
#        needed_pzs = True
#        needed_miss = False
#        needed_study = False
    elif department == 3:
        PATH = REPORTS_DIR+"СТ_{}_{}.xlsx".format(datetime.datetime.now().year, datetime.datetime.now().month)
        needed_losts = False
        needed_kcrs = False
        needed_exps = False
        needed_visions = False
        needed_pzs = False
        needed_miss = False
        needed_study = False
#    elif department == 5:
#        PATH = REPORTS_DIR+"КБ_{}_{}.xlsx".format(datetime.datetime.now().year, datetime.datetime.now().month)
#        needed_losts = False
#        needed_kcrs = False
#        needed_exps = False
#        needed_visions = False
#        needed_pzs = False
#        needed_miss = False
#        needed_study = False
#    elif department == 6:
#        PATH = REPORTS_DIR+"ЭЛЬБА_{}_{}.xlsx".format(datetime.datetime.now().year, datetime.datetime.now().month)
#        needed_losts = False
#        needed_kcrs = False
#        needed_exps = False
#        needed_visions = False
#        needed_pzs = False
#        needed_miss = False
#        needed_study = False
    try:
        wb = load_workbook(PATH)
    except:
        wb = Workbook()
    last_ws = wb[wb.sheetnames[-1]]
    last_date = last_ws.title
    today = str(datetime.date.today())

    if today == last_date:
        target = wb[wb.sheetnames[-1]]
        wb.remove(target)
        ws = wb.create_sheet(today)
    else:
        ws = wb.create_sheet(today)

    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    sql = """SELECT pings, knowledges, checks, dezh, miss, study, break_time, reboot, op_time, losts, exps, kcrs, visions, pzs, pings_canceled, dezh_canceled, date, total_time, total_dezh_time, hards
            FROM stats
            WHERE department={}""".format(department)
    res = mycursor.execute(sql)
    data = res.fetchall()
    data = data[0]
    # sql1 = """SELECT SUM(pings), SUM(knowledges), SUM(checks), SUM(dezh), SUM(losts), SUM(kcrs), SUM(exps)
    #         FROM stats
    #         WHERE department={}""".format(department)
    # res = mycursor.execute(sql1)
    # total = 0
    # for qty in res.fetchall()[0]:
    #     total += qty
    # s_sql = """SELECT SUM(dezh), SUM(study), SUM(break_time), SUM(reboot), SUM(op_time), SUM(miss)
    #         FROM staff
    #         WHERE department={}""".format(department)
    # res = mycursor.execute(s_sql)
    # total_dezh = 0
    # for qty in res.fetchall()[0]:
    #     total_dezh += qty
    pings, knowledges, checks, dezh, miss, study, break_time, reboot, op_time, losts, exps, kcrs, visions, pzs, pings_canceled, dezh_canceled, date, total_time, total_dezh_time, hards = data

    for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T"]:
        ws.column_dimensions[col].width = 15

    ws.append(["Всего вызовов", pings+knowledges+checks+dezh+miss+study+losts+kcrs+exps+visions+pzs+break_time+op_time+reboot])
    if pings > 0:
        ws.append(["Всего пингов", pings])
    if knowledges > 0:
        ws.append(["Всего привязок знаний", knowledges])
    if hards > 0:
        ws.append(["Всего сложных", hards])
    if checks > 0:
        ws.append(["Всего проверок", checks])
    if losts > 0 and needed_losts:
        ws.append(["Всего потеряшек", losts])
    if kcrs > 0 and needed_kcrs:
        ws.append(["Всего КЦР", kcrs])
    if exps > 0 and needed_exps:
        ws.append(["Всего эксперт", exps])
    if visions > 0 and needed_visions:
        ws.append(["Всего визий", visions])
    if pzs > 0 and needed_pzs:
        ws.append(["Всего проверок знаний", pzs])
    if dezh > 0:
        ws.append(["Всего дежурок", dezh])
    if break_time > 0:
        ws.append(["Всего пропусков обеда", break_time])
    if miss > 0 and needed_miss:
        ws.append(["Всего пропусков фикс.перерыва", miss])
    if study > 0 and needed_study:
        ws.append(["Всего обучениё/Прослушок", study])
    if reboot > 0:
        ws.append(["Всего перезагрузок", reboot])
    if op_time > 0:
        ws.append(["Всего ОП", op_time])
    ws.append(["Всего отмен", pings_canceled+dezh_canceled])
    if pings_canceled > 0:
        ws.append(["Всего пингов отменено", pings_canceled])
    if dezh_canceled > 0:
        ws.append(["Всего дежурок отменено", dezh_canceled])
    # if total == 0:
    #     ws.append(["Ожидание ответа пинга", "Неизвестно"])
    # else:
    #     ws.append(["Ожидание ответа пинга", "{}:{}".format((total_time/total)//60, (total_time/total)%60)])
    # if total_dezh == 0:
    #     ws.append(["Ожидание ответа дежурки", "Неизвестно"])
    # else:
    #     ws.append(["Ожидание ответа дежурки", "{}:{}".format((total_dezh_time/total_dezh)//60, (total_dezh_time/total_dezh)%60)])

    if department != 3:
        ws.append(["--------------"])
        ws.append(["СУПЕРВИЗОРСКАЯ"])
        ws.append(["--------------"])
        sql = """SELECT last_name, first_name, pings, knowledges, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, losts, kcrs, visions, pzs, miss
                FROM supers
                WHERE department={}
                ORDER BY last_name ASC""".format(department)
        mycursor.execute(sql)
        su_list = mycursor.fetchall()
        if department == 1:
            ws.append(["Фамилия","Имя","Пинги","Знания","Гл.деж","Вр.деж","Обуч./Прос(ДА)","Обуч./Прос(НЕТ)","Перерыв","Перезагрузка(ДА)","Перезагрузка(НЕТ)","ОП","Потеряхи","КЦР","Визия","Проверка знаний","Фикс.перерыв","Всего"])
    #    elif department == 2:
    #        ws.append(["Фамилия","Имя","Пинги","Знания","Гл.деж","Вр.деж","Визия","Проверка знаний","Всего"])
    #    elif department in [5,6]:
    #        ws.append(["Фамилия","Имя","Пинги","Гл.деж","Вр.деж","Всего"])
        for row in su_list:
            try:
                last_name, first_name, pings, knowledges, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, losts, kcrs, visions, pzs, miss = row
                summary = pings+knowledges+f_dezh+h_dezh+study_yes+study_no+break_time+reboot_yes+reboot_no+op_time+losts+kcrs+visions+pzs+miss
                if department == 1:
                    ws.append([last_name, first_name, pings, knowledges, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, losts, kcrs, visions, pzs, miss, summary])
        #        elif department == 2:
        #            ws.append([last_name, first_name, pings, f_dezh, h_dezh, visions, pzs, summary])
        #        elif department in [5,6]:
        #            ws.append([last_name, first_name, pings, f_dezh, h_dezh, summary])
            except:
                pass
        su_sql = """SELECT SUM(pings), SUM(knowledges), SUM(f_dezh), SUM(h_dezh), SUM(study_yes), SUM(study_no), SUM(break_time), SUM(reboot_yes), SUM(reboot_no), SUM(op_time), SUM(losts), SUM(kcrs), SUM(visions), SUM(pzs), SUM(miss)
                FROM supers
                WHERE department={}""".format(department)
        mycursor.execute(su_sql)
        su_data = mycursor.fetchall()[0]
        t_pings, t_knowledges, t_f_dezh, t_h_dezh, t_study_yes, t_study_no, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_losts, t_kcrs, t_visions, t_pzs, t_miss = su_data
        t_total = t_pings+t_knowledges+t_f_dezh+t_h_dezh+t_study_yes+t_study_no+t_break_time+t_reboot_yes+t_reboot_no+t_op_time+t_losts+t_kcrs+t_visions+t_pzs+t_miss
        if department == 1:
            ws.append(["----","ИТОГО", t_pings, t_knowledges, t_f_dezh, t_h_dezh, t_study_yes, t_study_no, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_losts, t_kcrs, t_visions, t_pzs, t_miss, t_total])
    #    elif department == 2:
    #        ws.append(["----","ИТОГО", t_pings, t_knowledges, t_f_dezh, t_h_dezh, t_visions, t_pzs, t_total])
    #    elif department in [5,6]:
    #        ws.append(["----","ИТОГО", t_pings, t_f_dezh, t_h_dezh, t_total])

    # if department != 3:
    #     ws.append(["--------------"])
    #     ws.append(["ПОМОГАТОРСКАЯ"])
    #     ws.append(["--------------"])
    #     sql = """SELECT last_name, first_name, pings, knowledges, checks, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, visions, pzs, hards_yes, hards_no
    #             FROM helpers 
    #             WHERE department={}
    #             ORDER BY last_name ASC""".format(department)
    #     mycursor.execute(sql)
    #     he_list = mycursor.fetchall()
    #     if department == 1:
    #         ws.append(["Фамилия","Имя","Пинги","Знания","Гл.деж","Вр.деж","Обуч./Прос(ДА)","Обуч./Прос(НЕТ)","Перерыв","Перезагрузка(ДА)","Перезагрузка(НЕТ)","ОП","Визия","Проверка знаний","Всего"])
    #     # elif department == 2:
    #     #     ws.append(["Фамилия","Имя","Пинги","Знания","Гл.деж","Вр.деж","Визия","Проверка знаний","Всего"])
    #     # elif department in [5,6]:
    #     #     ws.append(["Фамилия","Имя","Пинги","Гл.деж","Вр.деж","Всего"])
    #     for row in he_list:
    #         try:
    #             last_name, first_name, pings, knowledges, checks, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, visions, pzs, hards_yes, hards_no = row
    #             summary = pings+knowledges+checks+f_dezh+h_dezh+study_yes+study_no+break_time+reboot_yes+reboot_no+op_time+visions+pzs+hards_yes+hards_no
    #             if summary != 0:
    #                 if department == 1:
    #                     ws.append([last_name, first_name, pings, knowledges, f_dezh, h_dezh, study_yes, study_no, break_time, reboot_yes, reboot_no, op_time, visions, pzs, summary])
    #                 # elif department == 2:
    #                 #     ws.append([last_name, first_name, pings, knowledges, f_dezh, h_dezh, visions, pzs, summary])
    #                 # elif department in [5,6]:
    #                 #     ws.append([last_name, first_name, pings, f_dezh, h_dezh, summary])    
    #         except:
    #             pass
    #     h_sql = """SELECT SUM(pings), SUM(knowledges), SUM(checks), SUM(f_dezh), SUM(h_dezh), SUM(study_yes), SUM(study_no), SUM(break_time), SUM(reboot_yes), SUM(reboot_no), SUM(op_time), SUM(visions), SUM(pzs), sum(hards_yes), sum(hards_no)
    #                 FROM helpers 
    #                 WHERE department={}""".format(department)
    #     mycursor.execute(h_sql)
    #     h_data = mycursor.fetchall()[0]
    #     t_pings, t_knowledges, t_checks, t_f_dezh, t_h_dezh, t_study_yes, t_study_no, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_visions, t_pzs, t_hards_yes, t_hards_no = h_data
    #     t_total = t_pings+t_knowledges+t_checks+t_f_dezh+t_h_dezh+t_study_yes+t_study_no+t_break_time+t_reboot_yes+t_reboot_no+t_op_time+t_visions+t_pzs+t_hards_yes+t_hards_no
    #     if department == 1:
    #         ws.append(["----","ИТОГО",t_pings, t_knowledges, t_f_dezh, t_h_dezh, t_study_yes, t_study_no, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_visions, t_pzs, t_total])
    #     # elif department == 2:
    #     #     ws.append(["----","ИТОГО",t_pings, t_knowledges, t_f_dezh, t_h_dezh, t_visions, t_pzs, t_total])
    #     # elif department in [5,6]:
    #     #     ws.append(["----","ИТОГО",t_pings, t_f_dezh, t_h_dezh, t_total])

    else:
        ws.append(["--------------"])
        ws.append(["НАСТАВНИЧЕСКАЯ"])
        ws.append(["--------------"])
        sql = """SELECT last_name, first_name, pings, knowledges, checks, hards_yes, hards_no, f_dezh, break_time, reboot_yes, reboot_no, op_time, visions, pzs
                FROM mentors
                WHERE department={}
                ORDER BY last_name ASC""".format(department)
        mycursor.execute(sql)
        he_list = mycursor.fetchall()
        ws.append(["Фамилия","Имя","Пинги","Знания","Проверки","Сложный(ДА)","Сложный(НЕТ)","Дежурный","Перерыв","Перезагрузка(ДА)","Перезагрузка(НЕТ)","ОП","Визия","Всего"])
        for row in he_list:
            try:
                last_name, first_name, pings, knowledges, checks, hards_yes, hards_no, f_dezh, break_time, reboot_yes, reboot_no, op_time, visions = row
                summary = pings+knowledges+checks+f_dezh+break_time+reboot_yes+reboot_no+op_time+hards_yes+hards_no+visions
                if summary != 0:
                    ws.append([last_name, first_name, pings, knowledges, checks, hards_yes, hards_no, f_dezh, break_time, reboot_yes, reboot_no, op_time, visions, summary])
            except:
                pass
        h_sql = """SELECT SUM(pings), SUM(knowledges), SUM(checks), sum(hards_yes), sum(hards_no), SUM(f_dezh), SUM(break_time), SUM(reboot_yes), SUM(reboot_no), SUM(op_time), sum(visions)
                    FROM mentors
                    WHERE department={}""".format(department)
        mycursor.execute(h_sql)
        h_data = mycursor.fetchall()[0]
        t_pings, t_knowledges, t_checks, t_hards_yes, t_hards_no, t_f_dezh, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_visions = h_data
        t_total = t_pings+t_knowledges+t_checks+t_f_dezh+t_break_time+t_reboot_yes+t_reboot_no+t_op_time+t_hards_yes+t_hards_no+t_visions
        ws.append(["----","ИТОГО",t_pings, t_knowledges, t_checks, t_hards_yes, t_hards_no, t_f_dezh, t_break_time, t_reboot_yes, t_reboot_no, t_op_time, t_visions, t_total])

    ws.append(["--------------"])
    ws.append(["РАБОТЯГИ"])
    ws.append(["--------------"])
    sql = """SELECT adress, last_name, first_name, pings, knowledges, hards, checks, dezh, study, break_time, reboot, op_time, losts, kcrs, exps, miss, pings_canceled, dezh_canceled
            FROM staff
            WHERE department={}
            ORDER BY adress, last_name""".format(department)
    mycursor.execute(sql)
    st_list = mycursor.fetchall()
    if department == 1:
        ws.append(["Фамилия","Имя","Пинги","Знания","Пинг отмен.","Дежурка","Обуч./Прос.","Перерыв","Перезагрузка","ОП","Дежурка отмен.","Потеряхи","КЦР","Эксперты","Фикс.перерыв","Всего","Всего отмен."])
#    elif department == 2:
#        ws.append(["Фамилия","Имя","Пинги","Знания","Пинг отмен.","Дежурка","Перерыв","Перезагрузка","ОП","Дежурка отмен.","Эксперты","Всего","Всего отмен."])
    elif department == 3:
        ws.append(["Набор","Фамилия","Имя","Пинги","Знания","Сложный","Проверки","Пинг отмен.","Дежурка","Перерыв","Перезагрузка","ОП","Дежурка отмен.","Всего","Всего отмен."])
#    elif department in [5,6]:
#        ws.append(["Фамилия","Имя","Пинги","Пинг отмен.","Дежурка","Дежурка отмен.","Всего","Всего отмен."])
    for row in st_list:
        try:
            adress, last_name, first_name, pings, knowledges, hards, checks, dezh, study, break_time, reboot, op_time, losts, kcrs, exps, miss, pings_canceled, dezh_canceled = row
            summary = pings+knowledges+checks+dezh+study+break_time+reboot+op_time+losts+kcrs+exps+miss+hards
            canceled_summary = pings_canceled + dezh_canceled
            if summary != 0:
                if department == 1:
                    ws.append([last_name, first_name, pings, knowledges, pings_canceled, dezh, study, break_time, reboot, op_time, dezh_canceled, losts, kcrs, exps, miss, summary, canceled_summary])
#                elif department == 2:
#                    ws.append([last_name, first_name, pings, knowledges, pings_canceled, dezh, break_time, reboot, op_time, dezh_canceled, exps, summary, canceled_summary])
                elif department == 3:
                    ws.append([adress, last_name, first_name, pings, knowledges, hards, checks, pings_canceled, dezh, break_time, reboot, op_time, dezh_canceled, summary, canceled_summary])
#                elif department in [5,6]:
#                    ws.append([last_name, first_name, pings, pings_canceled, dezh, dezh_canceled, summary, canceled_summary])
        except:
            pass
    s_sql = """SELECT SUM(pings), SUM(knowledges), sum(hards), SUM(checks), SUM(dezh), SUM(study), SUM(break_time), SUM(reboot), SUM(op_time), SUM(losts), SUM(kcrs), SUM(exps), SUM(miss), SUM(pings_canceled), SUM(dezh_canceled)
            FROM staff
            WHERE department={}""".format(department)
    mycursor.execute(s_sql)
    s_data = mycursor.fetchall()[0]
    t_pings, t_knowledges, t_hards, t_checks, t_dezh, t_study, t_break_time, t_reboot, t_op_time, t_losts, t_kcrs, t_exps, t_miss, t_canceled_dezh, t_canceled_pings = s_data
    t_total = t_pings+t_knowledges+t_hards+t_checks+t_dezh+t_study+t_break_time+t_reboot+t_op_time+t_losts+t_kcrs+t_exps+t_miss
    t_canceled_total= t_canceled_dezh + t_canceled_pings
    if department == 1:
        ws.append(["----","ИТОГО",t_pings, t_knowledges, t_canceled_pings, t_dezh, t_study, t_break_time, t_reboot, t_op_time, t_canceled_dezh, t_losts, t_kcrs, t_exps, t_miss, t_total, t_canceled_total])
#    elif department == 2:
#        ws.append(["----","ИТОГО",t_pings, t_knowledges, t_canceled_pings, t_dezh, t_break_time, t_reboot, t_op_time, t_canceled_dezh, t_exps, t_total, t_canceled_total])
    elif department == 3:
        ws.append(["----","----","ИТОГО",t_pings, t_knowledges, t_hards, t_checks, t_canceled_pings, t_dezh, t_break_time, t_reboot, t_op_time, t_canceled_dezh, t_total, t_canceled_total])
#    elif department in [5,6]:
#        ws.append(["----","ИТОГО",t_pings, t_canceled_pings, t_dezh, t_canceled_dezh, t_total, t_canceled_total])
    wb.save(PATH)


def is_today(department):
    mydb = sqlite3.connect(DB_DIR)
    mycursor = mydb.cursor()
    today = datetime.date.today()
    sql = "SELECT date FROM stats WHERE department={}".format(department)
    mycursor.execute(sql)
    result = mycursor.fetchall()
    s_date = result[0][0].split("-")
    date2 = "{}-{:0>2}-{:0>2}".format(s_date[0], s_date[1], s_date[2])

    if str(today) != str(date2):
        if department == 1:
            make_statistic(1)
#        elif department == 2:
#            make_statistic(2)
        elif department == 3:
            make_statistic(3)
#        elif department == 5:
#            make_statistic(5)
#        elif department == 6:
#            make_statistic(6)

        sql = """UPDATE stats
                SET total_time=0, total_dezh_time=0, pings=0, knowledges=0, hards=0, checks=0, dezh=0, study=0, break_time=0, reboot=0, op_time=0, checks=0, losts=0, exps=0, kcrs=0, visions=0, pzs=0, pings_canceled=0, dezh_canceled=0, miss=0"""
        mycursor.execute(sql)

        sql = """UPDATE staff
                SET total_time=0, total_dezh_time=0, hards=0, time_to_answer=0, time_to_answer_dezh=0, last_ping_mid=-1, ping_type='empty', ping_status=0, last_dezh_mid=-1, dezh_status=0, pings_canceled=0, dezh_canceled=0, pings=0, knowledges=0, checks=0, dezh=0, study=0, break_time=0, reboot=0, op_time=0, checks=0, kcrs=0, losts=0, exps=0, miss=0"""
        mycursor.execute(sql)

        sql = """UPDATE helpers
                SET pings=0, knowledges=0, hards_yes=0, hards_no=0, checks=0, f_dezh=0, h_dezh=0, study_yes=0, study_no=0, break_time=0, reboot_yes=0, reboot_no=0, op_time=0, checks=0, kcrs=0, visions=0, pzs=0"""
        mycursor.execute(sql)

        sql = """UPDATE supers
                SET pings=0, knowledges=0, hards_yes=0, hards_no=0, checks=0, f_dezh=0, h_dezh=0, study_yes=0, study_no=0, break_time=0, reboot_yes=0, reboot_no=0, op_time=0, losts=0, exps=0, kcrs=0, visions=0, pzs=0, miss=0"""
        mycursor.execute(sql)

        sql = """UPDATE mentors
                SET pings=0, knowledges=0, hards_yes=0, hards_no=0, checks=0, f_dezh=0, break_time=0, reboot_yes=0, reboot_no=0, op_time=0"""
        mycursor.execute(sql)

        mydb.commit()
