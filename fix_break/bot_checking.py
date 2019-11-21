import datetime
from openpyxl import load_workbook
from pprint import pprint
import time

from telebot import types

from functions import make_statistic


def auto_alert(bot):
    """Автоматическая отправка уведомлений о смене перерыва"""
    sleep_time = 120  # На сколько отправлять в сон
    path = "\\\\ra-fs\\ra-joint\\Проекты ОА\\Эксперимент в УЦ\\"
    xlsfile_origname = "График перерывов УЦ.xlsx"  # Имя исходного файла
    xlsfile_origpath = path + xlsfile_origname  # Формируем полный путь до файла

    while True:
        current_time = datetime.datetime.today().strftime("%H:%M")  # Определяем актуальное время
        current_date = datetime.datetime.today().strftime("%d.%m")  # Определяем актуальную дату
        sheet_name = current_date  # Указываем имя листа = Сегодняшняя дата
        wb = load_workbook(xlsfile_origpath)  # Открываем книгу с оригинальными данными
        sheet = wb[sheet_name]  # Открываем лист с оригинальными данными
        start_row = 2
        start_col = 3
        end_row = sheet.max_row+1
        end_col = sheet.max_column+1
        # ========================== Списки и словари ==================================================================
        operator_list = {row: sheet.cell(row=row, column=2).value for row in range(start_row, end_row, 1)}
        breaktime_list = {column: sheet.cell(row=1, column=column).value for column in range(start_col, end_col, 1)}
        # ========================== Логика работы =====================================================================
        print("Начинаем проверять данные")
        for n_row, operator in operator_list.items():  # Перебераем операторов по порядку
            try:
                for n_col in range(start_col, end_col-1, 1):  # Проверяем по всем ячейкам
                    if n_col == 38:
                        break

                    data = sheet.cell(row=1, column=n_col+1).value
                    value = sheet.cell(row=n_row, column=n_col+1).value
                    table_time = data.split("-")[0]
                    
                    if value is not None:
                        if current_time <= table_time :
                            time_for_break = sheet.cell(row=n_row, column=n_col+1).value
                            hours_diff = datetime.datetime.strptime(table_time, "%H:%M").hour - datetime.datetime.strptime(current_time, "%H:%M").hour
                            minute_diff = datetime.datetime.strptime(table_time, "%H:%M").minute - datetime.datetime.strptime(current_time, "%H:%M").minute

                            if hours_diff == 0:

                                if 1 < minute_diff <= 3:
                                    bot.send_message(
                                            chat_id=operator,
                                            text="Через пару минут (в {}) у тебя перерыв на {} минут.".format(
                                                    table_time, time_for_break))
                                    break

                                elif 9 < minute_diff <= 11:
                                    bot.send_message(
                                            chat_id=operator,
                                            text="В {} тебе нужно выйти перерыв на {} минут.".format(
                                                    table_time, time_for_break))
                                    break

                            elif hours_diff == 1:

                                if -59 < minute_diff <= -57:
                                    bot.send_message(
                                            chat_id=operator,
                                            text="Через пару минут (в {}) у тебя перерыв на {} минут.".format(
                                                    table_time, time_for_break))
                                    break

                                elif -51 < minute_diff <= -49:
                                    bot.send_message(
                                            chat_id=operator,
                                            text="В {} тебе нужно выйти перерыв на {} минут.".format(
                                                    table_time, time_for_break))

                                    break

            except Exception as e:
                print(e)
        print("Устал, пошел спать")
        make_statistic(1)
        make_statistic(3)
        time.sleep(sleep_time)


def send_grafik(bot, uid):
    """График для конс-а"""
    path = "\\\\ra-fs\\ra-joint\\Проекты ОА\\Эксперимент в УЦ\\"
    xlsfile_origname = "График перерывов УЦ.xlsx"  # Имя исходного файла
    xlsfile_origpath = path + xlsfile_origname  # Формируем полный путь до файла
    current_date = datetime.datetime.today().strftime("%d.%m")  # Определяем актуальную дату
    sheet_name = current_date  # Указываем имя листа = Сегодняшняя дата
    wb = load_workbook(xlsfile_origpath)  # Открываем книгу с оригинальными данными
    sheet = wb[sheet_name]  # Открываем лист с оригинальными данными
    start_row = 2
    start_col = 3
    end_row = sheet.max_row+1
    end_col = sheet.max_column+1
    # Распечатывание графика для нужного консультанта
    operator_list = {row: sheet.cell(row=row, column=2).value for row in range(start_row, end_row, 1)}
    l = list()
    
    for n_row, operator in operator_list.items():
        if operator == uid:
            for n_col in range(start_col, end_col, 1):
                # if n_col == sheet.max_column+1:
                #     pass
                cell_break = sheet.cell(row=n_row, column=n_col).value
                cell_time = sheet.cell(row=1, column=n_col).value
                if cell_break is not None:
                    l.append("{} : перерыв {} минут".format(cell_time, cell_break))

    ms_text = "\n".join(map(str, l))
    if len(ms_text) == 0:
        ms_text = "Тебя нет в графике"
    bot.send_message(
            chat_id=uid,
            text=ms_text)
